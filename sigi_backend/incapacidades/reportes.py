"""
AUSENTRACK - Generacion de reportes de ausentismo
Excel con openpyxl y PDF con reportlab
"""

import io
from datetime import date
from django.http import HttpResponse
from .models import Incapacidad


def get_queryset_reporte(params):
    qs = Incapacidad.objects.select_related('colaborador', 'entidad_emisora').order_by('-fecha_inicio')
    if params.get('colaborador'):
        qs = qs.filter(colaborador=params['colaborador'])
    if params.get('tipo'):
        qs = qs.filter(tipo=params['tipo'])
    if params.get('estado'):
        qs = qs.filter(estado=params['estado'])
    if params.get('area'):
        qs = qs.filter(colaborador__area__icontains=params['area'])
    if params.get('fecha_desde'):
        qs = qs.filter(fecha_inicio__gte=params['fecha_desde'])
    if params.get('fecha_hasta'):
        qs = qs.filter(fecha_inicio__lte=params['fecha_hasta'])
    return qs


def exportar_excel(queryset):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return None, 'openpyxl no esta instalado.'

    wb = Workbook()
    ws = wb.active
    ws.title = 'Reporte Ausentismo'

    # Estilos
    color_header  = '1F3864'
    font_header   = Font(bold=True, color='FFFFFF', size=11)
    fill_header   = PatternFill(fill_type='solid', fgColor=color_header)
    fill_fila_par = PatternFill(fill_type='solid', fgColor='EBF5FB')
    borde = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC'),
    )

    # Titulo
    ws.merge_cells('A1:J1')
    ws['A1'] = 'AUSENTRACK - Reporte de Ausentismo'
    ws['A1'].font      = Font(bold=True, size=14, color=color_header)
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:J2')
    ws['A2'] = f'Generado el {date.today().strftime("%d/%m/%Y")}'
    ws['A2'].font      = Font(size=10, color='888888')
    ws['A2'].alignment = Alignment(horizontal='center')

    # Encabezados
    encabezados = [
        'Colaborador', 'Cedula', 'Area', 'Tipo', 'Fecha Inicio',
        'Fecha Fin', 'Dias', 'Responsable Pago', 'Estado', 'Entidad Emisora'
    ]
    anchos = [28, 16, 18, 22, 14, 14, 8, 18, 12, 22]

    for col, (enc, ancho) in enumerate(zip(encabezados, anchos), start=1):
        celda             = ws.cell(row=4, column=col, value=enc)
        celda.font        = font_header
        celda.fill        = fill_header
        celda.alignment   = Alignment(horizontal='center', vertical='center')
        celda.border      = borde
        ws.column_dimensions[celda.column_letter].width = ancho

    ws.row_dimensions[4].height = 22

    # Datos
    tipo_map  = dict(Incapacidad.TIPO_CHOICES)
    resp_map  = dict(Incapacidad.RESPONSABLE_CHOICES)
    estado_map = dict(Incapacidad.ESTADO_CHOICES)

    total_dias = 0
    for fila_num, inc in enumerate(queryset, start=5):
        es_par = fila_num % 2 == 0
        valores = [
            inc.colaborador.nombre,
            inc.colaborador.cedula,
            inc.colaborador.area or '',
            tipo_map.get(inc.tipo, inc.tipo),
            inc.fecha_inicio.strftime('%d/%m/%Y'),
            inc.fecha_fin.strftime('%d/%m/%Y'),
            inc.dias,
            resp_map.get(inc.responsable_pago, inc.responsable_pago),
            estado_map.get(inc.estado, inc.estado),
            inc.entidad_emisora.nombre if inc.entidad_emisora else '',
        ]
        total_dias += inc.dias
        for col, valor in enumerate(valores, start=1):
            celda           = ws.cell(row=fila_num, column=col, value=valor)
            celda.border    = borde
            celda.alignment = Alignment(vertical='center')
            if es_par:
                celda.fill = fill_fila_par

    # Totales
    fila_total = len(list(queryset)) + 5
    ws.cell(row=fila_total, column=6, value='TOTAL DIAS:').font = Font(bold=True)
    celda_total = ws.cell(row=fila_total, column=7, value=total_dias)
    celda_total.font = Font(bold=True, color=color_header)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer, None


def exportar_pdf(queryset):
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    except ImportError:
        return None, 'reportlab no esta instalado.'

    buffer = io.BytesIO()
    doc    = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                               leftMargin=1.5*cm, rightMargin=1.5*cm,
                               topMargin=1.5*cm, bottomMargin=1.5*cm)

    styles   = getSampleStyleSheet()
    color_hd = colors.HexColor('#1F3864')

    estilo_titulo = ParagraphStyle('titulo', parent=styles['Title'],
                                   textColor=color_hd, fontSize=16, spaceAfter=4)
    estilo_sub    = ParagraphStyle('sub', parent=styles['Normal'],
                                   textColor=colors.grey, fontSize=9, spaceAfter=12)

    elementos = [
        Paragraph('AUSENTRACK', estilo_titulo),
        Paragraph(f'Reporte de Ausentismo — {date.today().strftime("%d/%m/%Y")}', estilo_sub),
    ]

    tipo_map   = dict(Incapacidad.TIPO_CHOICES)
    resp_map   = dict(Incapacidad.RESPONSABLE_CHOICES)
    estado_map = dict(Incapacidad.ESTADO_CHOICES)

    encabezados = ['Colaborador', 'Cedula', 'Tipo', 'Inicio', 'Fin', 'Dias', 'Responsable', 'Estado']
    filas = [encabezados]
    total_dias = 0

    for inc in queryset:
        total_dias += inc.dias
        filas.append([
            inc.colaborador.nombre,
            inc.colaborador.cedula,
            tipo_map.get(inc.tipo, inc.tipo),
            inc.fecha_inicio.strftime('%d/%m/%Y'),
            inc.fecha_fin.strftime('%d/%m/%Y'),
            str(inc.dias),
            resp_map.get(inc.responsable_pago, inc.responsable_pago),
            estado_map.get(inc.estado, inc.estado),
        ])

    filas.append(['', '', '', '', 'Total dias:', str(total_dias), '', ''])

    tabla = Table(filas, colWidths=[5*cm, 3*cm, 4.5*cm, 2.5*cm, 2.5*cm, 1.8*cm, 3*cm, 2.5*cm])
    tabla.setStyle(TableStyle([
        ('BACKGROUND',  (0, 0), (-1, 0),  color_hd),
        ('TEXTCOLOR',   (0, 0), (-1, 0),  colors.white),
        ('FONTNAME',    (0, 0), (-1, 0),  'Helvetica-Bold'),
        ('FONTSIZE',    (0, 0), (-1, 0),  9),
        ('ALIGN',       (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN',      (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTSIZE',    (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#EBF5FB')]),
        ('GRID',        (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
        ('FONTNAME',    (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('ROWHEIGHT',   (0, 0), (-1, -1), 20),
    ]))

    elementos.append(tabla)
    doc.build(elementos)
    buffer.seek(0)
    return buffer, None
