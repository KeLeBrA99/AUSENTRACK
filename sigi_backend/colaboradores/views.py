"""
SIGI - Vistas de colaboradores
"""

from rest_framework import generics, status
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated

from authentication.permissions import EsAdmin, EsTalentoHumano, EsNomina
from authentication.views import registrar_auditoria
from .models import Empresa, Entidad, Colaborador
from .serializers import (
    EmpresaSerializer,
    EntidadSerializer,
    ColaboradorSerializer,
    ColaboradorListSerializer,
)


# ── EMPRESA ──────────────────────────────────────────────────────────────────

class ListarEmpresasView(generics.ListAPIView):
    """
    GET /api/colaboradores/empresas/
    """
    queryset           = Empresa.objects.filter(activo=True)
    serializer_class   = EmpresaSerializer
    permission_classes = [IsAuthenticated]


# ── ENTIDADES ─────────────────────────────────────────────────────────────────

class ListarEntidadesView(generics.ListAPIView):
    """
    GET /api/colaboradores/entidades/?tipo=EPS
    GET /api/colaboradores/entidades/?tipo=ARL
    """
    serializer_class   = EntidadSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        qs   = Entidad.objects.filter(activo=True)
        tipo = self.request.query_params.get('tipo')
        if tipo:
            qs = qs.filter(tipo=tipo)
        return qs.order_by('nombre')


class CrearEntidadView(generics.CreateAPIView):
    """
    POST /api/colaboradores/entidades/crear/
    Solo ADMIN.
    """
    serializer_class   = EntidadSerializer
    permission_classes = [IsAuthenticated, EsAdmin]

    def perform_create(self, serializer):
        entidad = serializer.save()
        registrar_auditoria(self.request.user, 'CREATE', 'entidad', entidad.id_entidad, request=self.request)


class DetalleEntidadView(generics.RetrieveUpdateDestroyAPIView):
    """
    GET/PUT/PATCH/DELETE /api/colaboradores/entidades/<id>/
    Solo ADMIN.
    """
    queryset           = Entidad.objects.all()
    serializer_class   = EntidadSerializer
    permission_classes = [IsAuthenticated, EsAdmin]
    lookup_field       = 'id_entidad'

    def perform_update(self, serializer):
        entidad = serializer.save()
        registrar_auditoria(self.request.user, 'UPDATE', 'entidad', entidad.id_entidad, request=self.request)

    def perform_destroy(self, instance):
        registrar_auditoria(self.request.user, 'DELETE', 'entidad', instance.id_entidad, request=self.request)
        instance.activo = False
        instance.save()


# ── COLABORADORES ─────────────────────────────────────────────────────────────

class ListarColaboradoresView(generics.ListAPIView):
    """
    GET /api/colaboradores/
    Soporta filtros: ?nombre=&cedula=&area=&activo=
    """
    serializer_class   = ColaboradorListSerializer
    permission_classes = [IsAuthenticated, EsNomina]

    def get_queryset(self):
        qs     = Colaborador.objects.select_related('empresa', 'eps', 'arl')
        nombre = self.request.query_params.get('nombre')
        cedula = self.request.query_params.get('cedula')
        area   = self.request.query_params.get('area')
        activo = self.request.query_params.get('activo')

        if nombre:
            qs = qs.filter(nombre__icontains=nombre)
        if cedula:
            qs = qs.filter(cedula__icontains=cedula)
        if area:
            qs = qs.filter(area__icontains=area)
        if activo is not None:
            qs = qs.filter(activo=activo.lower() == 'true')

        return qs.order_by('nombre')


class CrearColaboradorView(generics.CreateAPIView):
    """
    POST /api/colaboradores/crear/
    Solo ADMIN y TALENTO_HUMANO.
    """
    serializer_class   = ColaboradorSerializer
    permission_classes = [IsAuthenticated, EsTalentoHumano]

    def perform_create(self, serializer):
        colaborador = serializer.save()
        registrar_auditoria(
            self.request.user, 'CREATE', 'colaborador',
            colaborador.id_colaborador, request=self.request
        )


class DetalleColaboradorView(generics.RetrieveUpdateDestroyAPIView):
    """
    GET/PUT/PATCH/DELETE /api/colaboradores/<id>/
    """
    queryset     = Colaborador.objects.select_related('empresa', 'eps', 'arl')
    lookup_field = 'id_colaborador'

    def get_serializer_class(self):
        if self.request.method == 'GET':
            return ColaboradorSerializer
        return ColaboradorSerializer

    def get_permissions(self):
        if self.request.method == 'GET':
            return [IsAuthenticated(), EsNomina()]
        return [IsAuthenticated(), EsTalentoHumano()]

    def perform_update(self, serializer):
        colaborador = serializer.save()
        registrar_auditoria(
            self.request.user, 'UPDATE', 'colaborador',
            colaborador.id_colaborador, request=self.request
        )

    def perform_destroy(self, instance):
        registrar_auditoria(
            self.request.user, 'DELETE', 'colaborador',
            instance.id_colaborador, request=self.request
        )
        instance.activo = False
        instance.save()


class BuscarColaboradorView(APIView):
    """
    GET /api/colaboradores/buscar/?q=termino
    Busqueda rapida por nombre o cedula.
    """
    permission_classes = [IsAuthenticated, EsNomina]

    def get(self, request):
        q = request.query_params.get('q', '').strip()
        if len(q) < 2:
            return Response([])
        colaboradores = Colaborador.objects.filter(
            activo=True
        ).filter(
            models.Q(nombre__icontains=q) | models.Q(cedula__icontains=q)
        ).select_related('eps', 'arl')[:10]

        data = [
            {
                'id_colaborador': c.id_colaborador,
                'cedula':         c.cedula,
                'nombre':         c.nombre,
                'cargo':          c.cargo,
                'area':           c.area,
                'eps':            c.eps.id_entidad if c.eps else None,
                'eps_nombre':     c.eps.nombre if c.eps else None,
                'arl':            c.arl.id_entidad if c.arl else None,
                'arl_nombre':     c.arl.nombre if c.arl else None,
            }
            for c in colaboradores
        ]
        return Response(data)


from django.db import models as django_models
BuscarColaboradorView.get.__globals__['models'] = django_models


class ImportarColaboradoresView(APIView):
    """
    POST /api/colaboradores/importar/
    Importa colaboradores desde un archivo Excel.
    Columnas esperadas: cedula, nombre, cargo, area, fecha_ingreso, nit_empresa, nit_eps, nit_arl
    Solo ADMIN y TALENTO_HUMANO.
    """
    permission_classes = [IsAuthenticated, EsTalentoHumano]

    def post(self, request):
        try:
            import openpyxl
        except ImportError:
            return Response({'error': 'openpyxl no esta instalado.'}, status=500)

        archivo = request.FILES.get('archivo')
        if not archivo:
            return Response({'error': 'No se envio ningun archivo.'}, status=400)

        try:
            wb   = openpyxl.load_workbook(archivo)
            ws   = wb.active
            rows = list(ws.iter_rows(values_only=True))
        except Exception:
            return Response({'error': 'No se pudo leer el archivo. Verifica que sea un Excel valido (.xlsx).'}, status=400)

        if len(rows) < 2:
            return Response({'error': 'El archivo no tiene datos.'}, status=400)

        # Detectar encabezados (primera fila)
        encabezados = [str(c).strip().lower() if c else '' for c in rows[0]]

        def col(nombre):
            try:
                return encabezados.index(nombre)
            except ValueError:
                return None

        i_cedula        = col('cedula')
        i_nombre        = col('nombre')
        i_cargo         = col('cargo')
        i_area          = col('area')
        i_fecha_ingreso = col('fecha_ingreso')
        i_nit_empresa   = col('nit_empresa')
        i_nit_eps       = col('nit_eps')
        i_nit_arl       = col('nit_arl')

        if i_cedula is None or i_nombre is None:
            return Response({
                'error': 'El archivo debe tener al menos las columnas: cedula, nombre.'
            }, status=400)

        from colaboradores.models import Empresa, Entidad, Colaborador
        from datetime import datetime

        exitosos  = 0
        errores   = []

        for num_fila, fila in enumerate(rows[1:], start=2):
            try:
                cedula = str(fila[i_cedula]).strip() if fila[i_cedula] else None
                nombre = str(fila[i_nombre]).strip() if fila[i_nombre] else None

                if not cedula or not nombre:
                    errores.append(f'Fila {num_fila}: cedula y nombre son obligatorios.')
                    continue

                if Colaborador.objects.filter(cedula=cedula).exists():
                    errores.append(f'Fila {num_fila}: cedula {cedula} ya existe, se omitio.')
                    continue

                # Empresa
                empresa = None
                if i_nit_empresa is not None and fila[i_nit_empresa]:
                    empresa = Empresa.objects.filter(nit=str(fila[i_nit_empresa]).strip()).first()
                if not empresa:
                    empresa = Empresa.objects.first()

                # EPS
                eps = None
                if i_nit_eps is not None and fila[i_nit_eps]:
                    eps = Entidad.objects.filter(nit=str(fila[i_nit_eps]).strip(), tipo='EPS').first()

                # ARL
                arl = None
                if i_nit_arl is not None and fila[i_nit_arl]:
                    arl = Entidad.objects.filter(nit=str(fila[i_nit_arl]).strip(), tipo='ARL').first()

                # Fecha ingreso
                fecha_ingreso = None
                if i_fecha_ingreso is not None and fila[i_fecha_ingreso]:
                    val = fila[i_fecha_ingreso]
                    if isinstance(val, datetime):
                        fecha_ingreso = val.date()
                    else:
                        try:
                            fecha_ingreso = datetime.strptime(str(val).strip(), '%Y-%m-%d').date()
                        except Exception:
                            try:
                                fecha_ingreso = datetime.strptime(str(val).strip(), '%d/%m/%Y').date()
                            except Exception:
                                pass

                Colaborador.objects.create(
                    cedula        = cedula,
                    nombre        = nombre,
                    cargo         = str(fila[i_cargo]).strip()         if i_cargo is not None and fila[i_cargo] else '',
                    area          = str(fila[i_area]).strip()          if i_area  is not None and fila[i_area]  else '',
                    fecha_ingreso = fecha_ingreso,
                    empresa       = empresa,
                    eps           = eps,
                    arl           = arl,
                )
                exitosos += 1

            except Exception as e:
                errores.append(f'Fila {num_fila}: {str(e)}')

        from authentication.views import registrar_auditoria
        registrar_auditoria(
            request.user, 'CREATE', 'colaborador', 0,
            detalle=f'Importacion masiva: {exitosos} exitosos, {len(errores)} errores',
            request=request
        )

        return Response({
            'exitosos': exitosos,
            'errores':  errores,
            'mensaje':  f'Se importaron {exitosos} colaboradores correctamente.'
        })


class DescargarPlantillaView(APIView):
    """
    GET /api/colaboradores/plantilla/
    Genera y descarga la plantilla Excel para importacion masiva.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            import io
            from django.http import HttpResponse
        except ImportError:
            return Response({'error': 'openpyxl no esta instalado.'}, status=500)

        wb = Workbook()
        ws = wb.active
        ws.title = 'Colaboradores'

        color_azul  = '1F3864'
        color_verde = '1D6F42'
        borde = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC'),
        )

        # Titulo
        ws.merge_cells('A1:H1')
        ws['A1'] = 'AUSENTRACK - Plantilla de Importacion de Colaboradores'
        ws['A1'].font      = Font(bold=True, size=13, color='FFFFFF')
        ws['A1'].fill      = PatternFill(fill_type='solid', fgColor=color_azul)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 28

        # Subtitulo
        ws.merge_cells('A2:H2')
        ws['A2'] = 'Completa los datos a partir de la fila 6. Campos en VERDE son obligatorios.'
        ws['A2'].font      = Font(size=10, color='555555', italic=True)
        ws['A2'].fill      = PatternFill(fill_type='solid', fgColor='EBF5FB')
        ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[2].height = 18

        # Leyenda
        ws.merge_cells('A3:D3')
        ws['A3'] = '* Campo obligatorio (verde) | Campos opcionales (azul)'
        ws['A3'].font      = Font(size=9, color='666666', italic=True)
        ws['A3'].alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[3].height = 14

        # Fila tecnica (nombres de columna para el sistema)
        columnas = [
            ('cedula',        'Cedula',          True,  18),
            ('nombre',        'Nombre Completo', True,  28),
            ('cargo',         'Cargo',           False, 22),
            ('area',          'Area',            False, 18),
            ('fecha_ingreso', 'Fecha Ingreso',   False, 16),
            ('nit_empresa',   'NIT Empresa',     False, 18),
            ('nit_eps',       'NIT EPS',         False, 16),
            ('nit_arl',       'NIT ARL',         False, 16),
        ]

        for col_num, (key, label, requerido, ancho) in enumerate(columnas, start=1):
            letra = get_column_letter(col_num)
            # Fila 4: nombre tecnico (lo lee el backend)
            c4 = ws.cell(row=4, column=col_num, value=key)
            c4.font      = Font(bold=True, size=9, color='888888')
            c4.fill      = PatternFill(fill_type='solid', fgColor='F2F2F2')
            c4.alignment = Alignment(horizontal='center')
            c4.border    = borde
            # Fila 5: etiqueta visual
            c5 = ws.cell(row=5, column=col_num, value=label + (' *' if requerido else ''))
            c5.font      = Font(bold=True, size=11, color='FFFFFF')
            c5.fill      = PatternFill(fill_type='solid', fgColor=color_verde if requerido else color_azul)
            c5.alignment = Alignment(horizontal='center', vertical='center')
            c5.border    = borde
            ws.column_dimensions[letra].width = ancho

        ws.row_dimensions[4].height = 16
        ws.row_dimensions[5].height = 24

        # Filas de ejemplo
        ejemplos = [
            ('12345678', 'Juan Carlos Perez',    'Analista',    'Tecnologia',     '2024-01-15', '900123456-1', '', ''),
            ('87654321', 'Maria Fernanda Gomez', 'Coordinadora','Talento Humano', '2023-06-01', '900123456-1', '', ''),
        ]
        for fila_num, datos in enumerate(ejemplos, start=6):
            es_par = fila_num % 2 == 0
            for col_num, valor in enumerate(datos, start=1):
                celda           = ws.cell(row=fila_num, column=col_num, value=valor)
                celda.fill      = PatternFill(fill_type='solid', fgColor='EBF5FB' if es_par else 'FFFFFF')
                celda.border    = borde
                celda.alignment = Alignment(vertical='center')
            ws.row_dimensions[fila_num].height = 20

        # Hoja referencia
        ws2 = wb.create_sheet('Referencia')
        ws2.merge_cells('A1:C1')
        ws2['A1'] = 'GUIA DE CAMPOS'
        ws2['A1'].font      = Font(bold=True, size=12, color='FFFFFF')
        ws2['A1'].fill      = PatternFill(fill_type='solid', fgColor=color_azul)
        ws2['A1'].alignment = Alignment(horizontal='center')
        refs = [
            ('Campo',         'Descripcion',                           'Ejemplo'),
            ('cedula',        'Numero de cedula',                      '12345678'),
            ('nombre',        'Nombre completo',                       'Juan Perez'),
            ('cargo',         'Cargo o puesto',                        'Analista'),
            ('area',          'Area o departamento',                   'Sistemas'),
            ('fecha_ingreso', 'Fecha formato YYYY-MM-DD o DD/MM/YYYY', '2024-01-15'),
            ('nit_empresa',   'NIT de la empresa (debe existir en el sistema)', '900123456-1'),
            ('nit_eps',       'NIT de la EPS registrada en el sistema', ''),
            ('nit_arl',       'NIT de la ARL registrada en el sistema', ''),
        ]
        for i, (c, d, e) in enumerate(refs, start=2):
            ws2.cell(row=i, column=1, value=c).font = Font(bold=(i == 2))
            ws2.cell(row=i, column=2, value=d)
            ws2.cell(row=i, column=3, value=e)
            for col in range(1, 4):
                ws2.cell(row=i, column=col).border    = borde
                ws2.cell(row=i, column=col).alignment = Alignment(vertical='center')
            ws2.row_dimensions[i].height = 18
        ws2.column_dimensions['A'].width = 18
        ws2.column_dimensions['B'].width = 45
        ws2.column_dimensions['C'].width = 20

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="plantilla_colaboradores.xlsx"'
        return response